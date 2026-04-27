#!/usr/bin/env python3
"""Bulk loader for configuration.Choice values from an Excel workbook."""

import argparse
import os

import django
from django.db import transaction
from openpyxl import load_workbook


def iter_rows(path: str, sheet_name: str | None = None):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {header: idx for idx, header in enumerate(headers)}

    required = {"Category", "Internal Value", "Display Value"}
    missing = required - set(header_map)
    if missing:
        raise ValueError(f"Workbook is missing required columns: {', '.join(sorted(missing))}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        category = row[header_map["Category"]]
        internal = row[header_map["Internal Value"]]
        display = row[header_map["Display Value"]]
        description = row[header_map.get("Description", -1)] if "Description" in header_map else None

        if not category or not internal:
            continue

        yield str(category).strip(), str(internal).strip(), (str(display).strip() if display else str(internal).strip()), description


def load_choices(path: str, sheet: str | None, truncate: bool) -> tuple[int, int]:
    from configuration.models import Choice

    rows = list(iter_rows(path, sheet))
    if not rows:
        return 0, 0

    categories = sorted({category for category, *_ in rows})

    with transaction.atomic():
        if truncate:
            Choice.objects.filter(category__in=categories).delete()

        created = 0
        updated = 0

        for category, internal, display, description in rows:
            defaults = {"display_value": display}
            obj, is_created = Choice.objects.update_or_create(
                category=category,
                internal_value=internal,
                defaults=defaults,
            )
            if is_created:
                created += 1
            else:
                updated += 1

    return created, updated


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--settings",
        default="FleetManager.settings",
        help="Django settings module (default: %(default)s)",
    )
    parser.add_argument(
        "--file",
        default=os.path.join("utils", "data", "choice_seed_data.xlsx"),
        help="Path to Excel workbook containing Choice seed data",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional worksheet name (defaults to the active sheet)",
    )
    parser.add_argument(
        "--truncate",
        action="store_true",
        help="Delete existing Choice rows for the categories present in the workbook before loading",
    )

    args = parser.parse_args(argv)

    os.environ.setdefault("DJANGO_SETTINGS_MODULE", args.settings)
    django.setup()

    created, updated = load_choices(args.file, args.sheet, args.truncate)
    print(f"Choice load complete. Created: {created}, Updated: {updated}")


if __name__ == "__main__":
    main()
