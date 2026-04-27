from django.db.models import Sum
from decimal import Decimal
from django.urls import reverse
from django.db.models import Count
from pathlib import Path
from FleetManager import settings
from entity.models import Organization, Driver, Vehicle
from configuration.models import Choice
from .forms import ShipmentExpenseForm, DriverAdvanceAdminForm
from django.utils.safestring import mark_safe
from django.contrib import admin, messages
from django.utils.html import format_html
from .models import (
    ConsignmentGroup, Consignment, Shipment, ShipmentExpense,
    DriverAdvance, ShipmentStatus, Diesel
)
from django.db import models

from django.forms import ModelChoiceField
from django.http import HttpResponse
from django.template.loader import render_to_string
from io import BytesIO
from xhtml2pdf import pisa
import base64
from .admin_mixins import NavigationButtonMixin
from django.urls import reverse
from django.urls import path
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter


class ConsignmentInline(admin.TabularInline):
    """
    Inline admin for managing consignments within a ConsignmentGroup.
    Shows key details with links to edit individual consignments.
    """
    model = Consignment
    extra = 1
    fields = [

        'consignment_id', 'consignor', 'consignee',
        'weight',

    ]
    readonly_fields = ['consignment_id']

    def get_queryset(self, request):
        qs = super().get_queryset(request)

        return qs.select_related('consignor', 'consignee', )


@admin.register(ConsignmentGroup)
# class ConsignmentGroupAdmin(admin.ModelAdmin):
class ConsignmentGroupAdmin(NavigationButtonMixin, admin.ModelAdmin):
    """
    Admin interface for ConsignmentGroup model.
    Provides comprehensive management for batching consignments.
    """
    readonly_fields = ('group_id', 'created_at', 'updated_at', 'total_weight', 'total_amount',
                       'consignment_count_display', 'route_display')

    fieldsets = (
        ('Group Information', {
            'fields': (
                'group_id',
                'consignments'

            ),
            'description': 'Basic group configuration and classification.'
        }),

        ('Calculated Totals', {
            'fields': [
                ('total_weight', 'total_amount'),
                ('consignment_count_display',),
            ],
            'classes': ('collapse',),
            'description': 'Auto-calculated fields from associated consignments.'
        }),
        ('Address', {
            'fields': (
                # ('origins', 'destinations'),
                ('route_display',)

            ),
            'classes': ('collapse',),
        }),

        ('Metadata', {
            'fields': (
                ('created_at', 'updated_at'),
                'created_by',
            ),
            'classes': ('collapse',),
        })
    )

    # inlines = [ConsignmentInline]
    filter_horizontal = ['consignments']
    list_display = (
        'group_id', 'next_step_list_button', 'consignment_count_display',
        'total_weight', 'total_amount', 'planned_dispatch_date',)

    # list_filter = (
    #     'planned_dispatch_date',)

    search_fields = (
        'group_id',
        'consignments__consignor__organization_name',
        'consignments__consignee__organization_name'
    )

    date_hierarchy = 'created_at'

    actions = ['calculate_totals', ]

    def route_display(self, obj):
        routes = []
        for c in obj.consignments.all():
            consignor_html = (
                f"<span style='color:blue'>"
                f"{c.consignor.address_line_1}, {c.consignor.address_line_2}, "
                f"{c.consignor.city}, {c.consignor.district}, "
                f"{c.consignor.state}, {c.consignor.pincode}"
                f"</span>"
            )
            consignee_html = (
                f"<span style='color:green'>"
                f"{c.consignee.address_line_1}, {c.consignee.address_line_2}, "
                f"{c.consignee.city}, {c.consignee.district}, "
                f"{c.consignee.state}, {c.consignee.pincode}"
                f"</span>"
            )
            routes.append(
                f"<strong>{c.consignment_id}</strong>: {consignor_html} → {consignee_html}"
            )

        return mark_safe("<br>".join(routes))

    # def get_queryset(self, request):
    #     """Optimize queries to avoid N+1 problems"""
    #     qs = super().get_queryset(request)
    #     return qs.select_related(
    #     ).prefetch_related('consignments').annotate(
    #         consignment_count=Count('consignments')
    #     )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return (
            qs.prefetch_related(
                'consignments',
                'consignments__consignor',
                'consignments__consignee'
            )
            .annotate(consignment_count=Count('consignments'))
        )

    # @admin.display(description="Consignments")
    # def consignment_count_display(self, obj):
    #     """Display number of consignments with link to view them"""
    #     count = getattr(obj, 'consignment_count', obj.consignments.count())
    #     if count > 0:
    #         url = reverse('admin:operations_consignment_changelist')
    #         return format_html(
    #             '<a href="{}?consignment_group__id__exact={}">{} consignment{}</a>',
    #             url, obj.pk, count, 's' if count != 1 else ''
    #         )
    #     return f"{count} consignments"

    @admin.display(description="Consignments")
    def consignment_count_display(self, obj):
        """Display number of consignments with link to view them"""
        count = getattr(obj, 'consignment_count', obj.consignments.count())

        if count > 0:
            url = reverse('admin:operations_consignment_changelist')
            return format_html(
                '<a href="{}?consignmentgroup__id__exact={}">{} consignment{}</a>',
                url,
                obj.pk,
                count,
                's' if count != 1 else ''
            )

        return f"{count} consignments"

    def calculate_totals(self, request, queryset):
        """Action to recalculate totals for selected groups"""
        for group in queryset:
            group.calculate_totals()
        count = queryset.count()
        self.message_user(request, f"Recalculated totals for {count} consignment groups.")

    def mark_as_active(self, request, queryset):
        """Mark selected groups as active"""
        updated = queryset.update(is_active=True)
        self.message_user(request, f"Marked {updated} groups as active.")

    def mark_as_inactive(self, request, queryset):
        """Mark selected groups as inactive"""
        updated = queryset.update(is_active=False)
        self.message_user(request, f"Marked {updated} groups as inactive.")

    calculate_totals.short_description = "Recalculate totals for selected groups"
    mark_as_active.short_description = "Mark selected groups as active"
    mark_as_inactive.short_description = "Mark selected groups as inactive"

    change_form_template = "admin/consignment/change_form.html"

    class Media:
        css = {
            "all": ("operations/css/navigation_buttons.css",)
        }

    @admin.display(description="Shipment", ordering=None)
    def next_step_list_button(self, obj):
        shipment = Shipment.objects.filter(consignment_group=obj).first()

        if shipment:
            return self.nav_button(
                "View",
                "admin:operations_shipment_change",
                shipment.pk
            )
        else:
            return self.nav_button(
                "Create",
                "admin:operations_shipment_add",
                params={"consignment_group": obj.pk}
            )

    def change_view(self, request, object_id, form_url='', extra_context=None):
        obj = self.get_object(request, object_id)

        shipment = Shipment.objects.filter(consignment_group=obj).first()

        # NEXT
        if shipment:
            next_step_url = reverse("admin:operations_shipment_change", args=[shipment.pk])
            next_step_label = "View Shipment"
        else:
            next_step_url = reverse("admin:operations_shipment_add") + f"?consignment_group={obj.pk}"
            next_step_label = "Create Shipment"

        # BACK → First Consignment
        first_consignment = obj.consignments.first()
        if first_consignment:
            back_step_url = reverse("admin:operations_consignment_change", args=[first_consignment.pk])
            back_step_label = "Back to Consignment"
        else:
            back_step_url = reverse("admin:index")
            back_step_label = "Back to Dashboard"

        extra_context = extra_context or {}
        extra_context.update({
            "next_step_url": next_step_url,
            "next_step_label": next_step_label,
            "back_step_url": back_step_url,
            "back_step_label": back_step_label,
        })

        return super().change_view(request, object_id, form_url, extra_context)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)

        # Auto-select consignment when coming from ConsignmentAdmin
        consignment_id = request.GET.get("consignment")

        if consignment_id and not obj:
            try:
                form.base_fields["consignments"].initial = [int(consignment_id)]
            except:
                pass

        return form

    # def get_form(self, request, obj=None, **kwargs):
    #     form = super().get_form(request, obj, **kwargs)
    #     self._selected_consignment = request.GET.get("consignment")
    #     return form

    def formfield_for_manytomany(self, db_field, request, **kwargs):
        if db_field.name == "consignments":
            # Get current object ID (None when adding)
            obj_id = request.resolver_match.kwargs.get("object_id")

            if obj_id:
                # Editing existing group → show:
                # 1. Consignments already in this group
                # 2. Consignments not assigned to any group
                kwargs["queryset"] = Consignment.objects.filter(
                    models.Q(consignmentgroup__isnull=True) |
                    models.Q(consignmentgroup__id=obj_id)
                ).distinct().order_by("-created_at")

            else:
                # Adding new group → show only unassigned consignments
                kwargs["queryset"] = Consignment.objects.filter(
                    consignmentgroup__isnull=True
                ).order_by("-created_at")

        return super().formfield_for_manytomany(db_field, request, **kwargs)


@admin.register(Consignment)
# class ConsignmentAdmin(admin.ModelAdmin):
class ConsignmentAdmin(NavigationButtonMixin, admin.ModelAdmin):
    """
    Updated admin interface for the new Consignment model.
    Handles complete consignor/consignee information directly.
    """

    readonly_fields = ("consignment_id", "created_at", "updated_at", "total_freight", 'route_summary_display')
    autocomplete_fields = ("consignor", "consignee", "from_location", "to_location", "vehicle_type", "material_type",
                           "packaging_type", "weight_unit", "freight_mode")
    fieldsets = (
        ('Basic Information', {
            'fields': (
                'consignment_id',
                ('consignor', 'consignee',),
                ('from_location', 'to_location',),
                'vehicle_type',
                'route_summary_display'

            ),
        }),

        ('Goods Information', {
            'fields': [
                ('material_type',),
                ('weight', 'weight_unit', 'volume'),
                ('number_of_packages', 'packaging_type'),
            ],

        }),
        ('Freight & Pricing', {
            'fields': [

                ('freight_mode', 'rate', 'total_freight'),
            ],
        }),

        ('Scheduling', {
            'fields': [
                ('schedule_date', 'scheduled_pickup_time'),
                ('expected_delivery_date', 'expected_delivery_time'),
            ],
            'classes': ('collapse',)
        }),

        ('Status & Metadata', {
            'fields': (
                ('created_at', 'updated_at'),
                'created_by',
            ),
            'classes': ('collapse',),
        })
        # ('Next Step', {
        #     'fields': ('next_step_buttons',),
        # }),
    )

    list_display = (

        'consignment_id', 'next_step_list_button', 'consignor',
        'consignee', 'freight_mode', 'rate', 'weight',
        'total_freight', 'schedule_date'
    )

    list_filter = (

    )
    # 'consignee',
    # 'consignor', 'created_at'
    search_fields = (
        'consignment_id',
        'consignor__organization_name',
        'consignee__organization_name',
        # "material_type__name",
        "from_location__name",
        "to_location__name",
    )

    date_hierarchy = 'created_at'

    actions = ['calculate_freight', ]

    class Meta:
        from django.db import models
        constraints = [
            models.CheckConstraint(
                check=~models.Q(from_location=models.F("to_location")),
                name="from_to_location_not_same"
            ),
        ]

    def get_queryset(self, request):
        qs = super().get_queryset(request)

        consignor = request.GET.get("consignor")
        consignee = request.GET.get("consignee")

        if consignor:
            qs = qs.filter(consignor_id=consignor)

        if consignee:
            qs = qs.filter(consignee_id=consignee)

        return qs.select_related(
            'consignor', 'consignee',
            'freight_mode', 'weight_unit'
        )

    def calculate_freight(self, request, queryset):
        """Action to recalculate freight for selected consignments"""
        for consignment in queryset:
            # This will trigger the calculation and update
            consignment.save()
        count = queryset.count()
        self.message_user(request, f"Recalculated freight for {count} consignments.")

    def mark_special_handling(self, request, queryset):
        """Mark selected consignments as requiring special handling"""
        updated = queryset.update(requires_special_handling=True)
        self.message_user(request, f"Marked {updated} consignments as requiring special handling.")

    def route_summary_display(self, obj):
        routes = []

        consignor_html = (
            f"<span style='color:blue'>"
            f"{obj.consignor.address_line_1}, {obj.consignor.address_line_2}, "
            f"{obj.consignor.city}, {obj.consignor.district}, "
            f"{obj.consignor.state}, {obj.consignor.pincode}"
            f"</span>"
        )
        consignee_html = (
            f"<span style='color:green'>"
            f"{obj.consignee.address_line_1}, {obj.consignee.address_line_2}, "
            f"{obj.consignee.city}, {obj.consignee.district}, "
            f"{obj.consignee.state}, {obj.consignee.pincode}"
            f"</span>"
        )
        routes.append(
            f"<strong>{obj.consignment_id}</strong>: {consignor_html} → {consignee_html}"
        )

        return mark_safe("<br>".join(routes))

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name in ('consignor', 'consignee'):
            kwargs['queryset'] = Organization.objects.filter(
                organization_type__internal_value='CONSIGNOR-AND-CONSIGNEE').order_by('organization_name')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    calculate_freight.short_description = "Recalculate freight for selected consignments"
    mark_special_handling.short_description = "Mark as requiring special handling"

    change_form_template = "admin/consignment/change_form.html"

    class Media:
        css = {
            "all": ("operations/css/navigation_buttons.css",)
        }

    # @admin.display(description="Next Step")
    # def next_step_buttons(self, obj):
    #     group = obj.consignmentgroup_set.first()
    #
    #     if group:
    #         return self.nav_button(
    #             "View Consignment Group",
    #             "admin:operations_consignmentgroup_change",
    #             group.pk
    #         )
    #     else:
    #         return self.nav_button(
    #             "Add to Consignment Group",
    #             "admin:operations_consignmentgroup_add",
    #             params={"consignment": obj.pk}
    #         )

    from django.utils.safestring import mark_safe

    # from django.urls import reverse, NoReverseMatch
    # print("URL:", reverse("admin:operations_consignmentgroup_change", args=[1]))

    @admin.display(description="Group", ordering=None)
    def next_step_list_button(self, obj):
        group = obj.consignmentgroup_set.first()

        if group:
            return self.nav_button("View", "admin:operations_consignmentgroup_change", group.pk)
        else:
            return self.nav_button("Add", "admin:operations_consignmentgroup_add", params={"consignment": obj.pk})

    def change_view(self, request, object_id, form_url='', extra_context=None):
        obj = self.get_object(request, object_id)

        group = obj.consignmentgroup_set.first()

        # NEXT
        if group:
            next_step_url = reverse("admin:operations_consignmentgroup_change", args=[group.pk])
            next_step_label = "View Consignment Group"
        else:
            next_step_url = reverse("admin:operations_consignmentgroup_add") + f"?consignment={obj.pk}"
            next_step_label = "Add to Consignment Group"

        # BACK → Dashboard
        back_step_url = reverse("admin:index")
        back_step_label = "Back to Dashboard"

        extra_context = extra_context or {}
        extra_context.update({
            "next_step_url": next_step_url,
            "next_step_label": next_step_label,
            "back_step_url": back_step_url,
            "back_step_label": back_step_label,
        })

        return super().change_view(request, object_id, form_url, extra_context)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}

        extra_context["consignors"] = (
            Organization.objects
            .filter(organization_type__internal_value='CONSIGNOR-AND-CONSIGNEE')
            .only("id", "organization_name")
            .order_by("organization_name")
        )

        extra_context["consignees"] = (
            Organization.objects
            .filter(organization_type__internal_value='CONSIGNOR-AND-CONSIGNEE')
            .only("id", "organization_name")
            .order_by("organization_name")
        )

        return super().changelist_view(request, extra_context)


class ShipmentStatusInline(admin.TabularInline):
    """Inline for tracking shipment status changes"""
    model = ShipmentStatus
    extra = 0
    fields = ('status', 'shipment_doc_type', 'shipment_document', 'updated_by', 'notes', 'effective_date')

    # readonly_fields = ('effective_date',)

    def get_formset(self, request, obj=None, **kwargs):
        """Auto-populate updated_by for new records"""
        FormSet = super().get_formset(request, obj, **kwargs)

        class CustomFormSet(FormSet):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                for form in self.forms:
                    if not form.instance.pk and hasattr(request, 'user'):
                        form.initial[
                            'updated_by'] = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username

        return CustomFormSet


class ShipmentExpenseInlineForShipment(admin.TabularInline):
    """Inline for managing shipment expenses"""
    model = ShipmentExpense
    form = ShipmentExpenseForm
    extra = 1
    fields = ('expense_by_combined', 'expense_type', 'amount', 'expense_date', 'description')
    verbose_name = "Shipment Expense"
    verbose_name_plural = "Shipment Expenses"

    def get_formset(self, request, obj=None, **kwargs):
        from datetime import date
        formset = super().get_formset(request, obj, **kwargs)
        if 'shipment' in formset.form.base_fields:
            del formset.form.base_fields['shipment']
        if 'expense_date' in formset.form.base_fields:
            formset.form.base_fields['expense_date'].initial = date.today()
        return formset


class DriverAdvanceInline(admin.TabularInline):
    """Inline for managing driver advances"""
    model = DriverAdvance
    extra = 0
    readonly_fields = ('date', 'remaining_balance')
    fields = ('driver', 'amount', 'description', 'is_settled', 'carried_forward', 'date', 'remaining_balance')


class DriverChoiceField(ModelChoiceField):
    def label_from_instance(self, obj):
        return f"{obj.id} - {obj.first_name} {obj.middle_name} {obj.last_name}".strip()


class ShipmentCompletionFilter(admin.SimpleListFilter):
    title = "Completion Status"
    parameter_name = "completion"

    def lookups(self, request, model_admin):
        return [
            ("pending", "Pending / In‑Progress"),
            ("completed", "Completed"),
        ]

    def queryset(self, request, queryset):
        from django.db.models import Exists, OuterRef
        from operations.models import ShipmentStatus

        completed_status = ShipmentStatus.objects.filter(
            shipment=OuterRef("pk"),
            status__internal_value__startswith="07_"
        )

        queryset = queryset.annotate(
            is_completed=Exists(completed_status)
        )

        if self.value() == "completed":
            return queryset.filter(is_completed=True)

        if self.value() == "pending":
            return queryset.filter(is_completed=False)

        return queryset


@admin.register(Shipment)
# class ShipmentAdmin(admin.ModelAdmin):
class ShipmentAdmin(NavigationButtonMixin, admin.ModelAdmin):
    """
    Updated admin interface for Shipment model working with ConsignmentGroup.
    """

    readonly_fields = (
        'shipment_id',
        'total_freight_amount',
        'invoice_link',
        'consignment_count_display',
        'route_summary_display',
        'total_distance',
        'generate_lr_button',
        # 'generate_invoice_no_button',
    )

    class Meta:
        indexes = [
            models.Index(fields=["shipment_id"]),
            models.Index(fields=["created_at"]),
            models.Index(fields=["vehicle"]),
            models.Index(fields=["driver"]),
        ]

    autocomplete_fields = (
        "consignment_group",
        "vehicle",
        "driver",
        "co_driver",
        "transporter",
        "broker",
        "planned_route",
        "actual_route",
    )
    fieldsets = (
        ('Shipment Assignment', {
            'fields': (
                'shipment_id',
                ('lr_no', 'generate_lr_button'),
                'consignment_group',
                'consignment_count_display',
                ('invoice_no',),
                'e_way_bill',
            ),
            'description': 'Basic shipment information and assigned consignment group.'
        }),
        ('Vehicle & Crew Assignment', {
            'fields': (
                ('vehicle', 'driver', 'co_driver'),
                ('transporter', 'broker'),
            ),
            'description': 'Assign vehicle, crew, and business partners for this shipment.'

        }),
        ('Schedule & Timing', {
            'fields': (
                ('planned_departure', 'actual_departure'),
                ('planned_arrival', 'actual_arrival'),
            ),
        }),
        ('Vehicle Tracking', {
            'fields': (
                ('odometer_start', 'odometer_end', 'total_distance'),
            ),
            'classes': ('collapse',)
        }),
        ('Route Information', {
            'fields': (
                'planned_route',
                'actual_route',
                'route_summary_display',
            ),
            'classes': ('collapse',)

        }),
        ('Financial Summary', {
            'fields': (
                ('freight_advance', 'total_freight_amount'),

                # ('total_expenses', 'balance'),
            ),
            'description': 'Financial overview. Add expenses in the "Shipment Expenses" section below.'
        }),
        ('Notes', {
            'fields': (

                'notes',
            ),
            # 'classes': ('collapse',)

        }),
        ('Metadata', {
            'fields': (
                'created_by',
            ),
            'classes': ('collapse',)
        })
    )

    class Media:
        js = ("operations/js/shipment_lr_generator.js",)

    # list_display = (
    #      'shipment_id', 'next_step_list_button','invoice_link','status_step_list_button','expense_step_list_button', 'vehicle', 'driver', 'consignment_group', 'consignment_count_display',
    #     'planned_departure')
    list_display = (
        'shipment_id', 'next_step_list_button', 'invoice_link', 'status_step_list_button', 'expense_step_list_button',
        'vehicle', 'driver', 'latest_status_badge', 'planned_departure')

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:shipment_id>/invoice/",
                self.admin_site.admin_view(self.generate_invoice_view),
                name="operations_shipment_invoice",
            ),
            path(
                "generate-lr-preview/",
                self.admin_site.admin_view(self.generate_lr_preview_view),
                name="operations_shipment_generate_lr_preview",
            ),
            # path(
            #     "generate-invoice-no-preview/",
            #     self.admin_site.admin_view(self.generate_invoice_no_preview_view),
            #     name="operations_shipment_generate_invoice_no_preview",
            # ),
        ]
        return custom_urls + urls

    @admin.display(description="LR")
    def generate_lr_button(self, obj):
        preview_url = reverse("admin:operations_shipment_generate_lr_preview")
        return format_html(
            '<button type="button" class="button" id="generate-lr-btn" '
            'data-preview-url="{}">Generate LR No</button>',
            preview_url
        )

    def generate_lr_preview_view(self, request):
        shipment_id = request.GET.get("shipment_id")
        exclude_pk = None

        if shipment_id and shipment_id.isdigit():
            exclude_pk = int(shipment_id)

        next_lr = Shipment.get_next_lr_no(exclude_pk=exclude_pk)
        return JsonResponse({"lr_no": next_lr})

    # @admin.display(description="LR")
    # def generate_invoice_no_button(self, obj):
    #     preview_url = reverse("admin:operations_shipment_generate_invoice_no_preview")
    #     return format_html(
    #         '<button type="button" class="button" id="generate-invoice-no-btn" '
    #         'data-preview-url="{}">Generate Invoice No</button>',
    #         preview_url
    #     )

    # def generate_invoice_no_preview_view(self, request):
    #     shipment_id = request.GET.get("shipment_id")
    #     exclude_pk = None
    #
    #     if shipment_id and shipment_id.isdigit():
    #         exclude_pk = int(shipment_id)
    #
    #     next_invoice_no = Shipment.get_next_invoice_no(exclude_pk=exclude_pk)
    #     return JsonResponse({"invoice_no": next_invoice_no})
    # @admin.display(description="Invoice")
    # def invoice_link(self, obj):
    #     url = reverse("admin:operations_shipment_invoice", args=[obj.pk])
    #     return format_html('<a target="_blank" href="{}">Generate PDF</a>', url)

    @admin.display(description="LR")
    def invoice_link(self, obj):
        url = reverse("admin:operations_shipment_invoice", args=[obj.pk])
        return format_html(
            '<a href="{}" target="_blank" class="button" '
            'style="margin-left: 5px;">LR</a>',
            url
        )

    def generate_invoice_view(self, request, shipment_id):
        shipment = Shipment.objects.select_related(
            "consignment_group",
            "vehicle",
            "driver",
        ).prefetch_related(
            "consignment_group__consignments__consignor",
            "consignment_group__consignments__consignee",
            "consignment_group__consignments__from_location",
            "consignment_group__consignments__to_location",
            "consignment_group__consignments__material_type",
            "consignment_group__consignments__weight_unit",
        ).get(pk=shipment_id)

        consignments = shipment.consignment_group.consignments.all() if shipment.consignment_group else []

        total_amount = shipment.total_freight_amount or Decimal("0.00")
        advance = shipment.freight_advance or Decimal("0.00")
        to_pay = total_amount - advance

        signature_uri = None
        signature_path = Path(settings.BASE_DIR) / "operations" / "static" / "images" / "authorized_signature.png"

        if signature_path.exists():
            with open(signature_path, "rb") as f:
                signature_uri = "data:image/png;base64," + base64.b64encode(f.read()).decode("utf-8")

        context = {
            "shipment": shipment,
            "consignments": consignments,
            "total_amount": total_amount,
            "advance": advance,
            "to_pay": to_pay,
            "signature_uri": signature_uri,
        }

        html = render_to_string("admin/shipments/shipment_invoice.html", context)

        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("utf-8")), result)

        if pdf.err:
            return HttpResponse("Error generating PDF", status=500)

        response = HttpResponse(result.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="shipment_invoice_{shipment.shipment_id}.pdf"'
        return response

    list_filter = ()

    # list_filter = (
    #     ShipmentCompletionFilter,'created_at')
    # 'vehicle__brand_name', 'transporter', 'broker',
    #     'planned_departure', 'actual_departure', ''
    #

    # search_fields = (
    #     'shipment_id', 'consignment_group__group_id', 'consignment_group__group_name',
    #     'vehicle__registration_number', 'driver__first_name', 'driver__last_name',
    #     'consignment_group__consignments__consignor_organization__organization_name'
    # )

    search_fields = (
        "shipment_id",
        "consignment_group__group_id",
        "vehicle__registration_number",
        "driver__first_name",
        "driver__last_name",
    )

    date_hierarchy = 'created_at'
    actions = ['calculate_totals', 'calculate_distances']

    # def get_queryset(self, request):
    #     qs = super().get_queryset(request)
    #
    #     # Add completion annotation here so filters can use it
    #     from django.db.models import Exists, OuterRef
    #     completed_status = ShipmentStatus.objects.filter(
    #         shipment=OuterRef("pk"),
    #         status__internal_value__startswith="07_"
    #     )
    #
    #     qs = qs.annotate(
    #         is_completed=Exists(completed_status)
    #     )
    #
    #     return qs.select_related(
    #         'consignment_group', 'vehicle', 'driver', 'transporter', 'broker'
    #     ).prefetch_related(
    #         'consignment_group__consignments', 'vehicle', 'driver'
    #     )

    def get_queryset(self, request):
        qs = super().get_queryset(request)

        # ---------------------------------------------------------
        # READ FILTER PARAMS
        # ---------------------------------------------------------
        vehicle = request.GET.get("vehicle") or None
        driver = request.GET.get("driver") or None
        status = request.GET.get("status") or None

        # ---------------------------------------------------------
        # APPLY FILTERS
        # ---------------------------------------------------------
        if vehicle:
            qs = qs.filter(vehicle_id=vehicle)

        if driver:
            qs = qs.filter(driver_id=driver)

        if status:
            qs = qs.filter(status_logs__status__internal_value=status)

        # ---------------------------------------------------------
        # RETURN WITH OPTIMIZED PREFETCH
        # ---------------------------------------------------------
        return qs.select_related(
            'consignment_group',
            'vehicle',
            'driver',
            'transporter',
            'broker'
        ).prefetch_related(
            'status_logs__status',
            'consignment_group__consignments',
            'vehicle',
            'driver'
        )

    @admin.display(description="Consignments")
    def consignment_count_display(self, obj):
        if not obj.consignment_group:
            return "No group assigned"

        consignments = list(obj.consignment_group.consignments.all())
        count = len(consignments)

        if count == 1:
            consignment = consignments[0]
            url = reverse("admin:operations_consignment_change", args=[consignment.pk])
            return format_html('<a href="{}">{} consignment</a>', url, count)

        url = reverse("admin:operations_consignment_changelist")
        return format_html(
            '<a href="{}?consignment_group__id__exact={}">{} consignments</a>',
            url,
            obj.consignment_group.pk,
            count
        )

    def route_summary_display(self, obj):
        routes = []
        for c in obj.consignment_group.consignments.all():
            consignor_html = (
                f"<span style='color:blue'>"
                f"{c.consignor.address_line_1}, {c.consignor.address_line_2}, "
                f"{c.consignor.city}, {c.consignor.district}, "
                f"{c.consignor.state}, {c.consignor.pincode}"
                f"</span>"
            )
            consignee_html = (
                f"<span style='color:green'>"
                f"{c.consignee.address_line_1}, {c.consignee.address_line_2}, "
                f"{c.consignee.city}, {c.consignee.district}, "
                f"{c.consignee.state}, {c.consignee.pincode}"
                f"</span>"
            )
            routes.append(
                f"<strong>{c.consignment_id}</strong>: {consignor_html} → {consignee_html}"
            )

        return mark_safe("<br>".join(routes))

    def calculate_totals(self, request, queryset):
        """Action to recalculate totals for selected shipments"""
        for shipment in queryset:
            shipment.calculate_totals()
        count = queryset.count()
        self.message_user(request, f"Recalculated totals for {count} shipments.")

    def calculate_distances(self, request, queryset):
        """Action to calculate distances from odometer readings"""
        count = 0
        for shipment in queryset:
            if shipment.calculate_distance():
                count += 1
        self.message_user(request, f"Calculated distances for {count} shipments.")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name in ("driver", "co_driver"):
            kwargs["queryset"] = Driver.objects.filter(is_active=True)
        elif db_field.name == "vehicle":
            kwargs["queryset"] = Vehicle.objects.filter(is_active=True)

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # change_form_template = "admin/consignment/change_form.html"
    change_form_template = "admin/operations/shipment/change_form.html"

    # change_list_template = "admin/consignment/change_list.html"
    change_list_template = "admin/operations/shipment/change_list.html"

    class Media:
        css = {
            "all": ("operations/css/navigation_buttons.css",)
        }

    @admin.display(description="Invoice", ordering=None)
    def next_step_list_button(self, obj):
        invoice = obj.invoices.first()  # related_name='invoices'

        if invoice:
            return self.nav_button(
                "View",
                "admin:financial_invoice_change",  # FIXED
                invoice.pk
            )
        else:
            return self.nav_button(
                "Create",
                "admin:financial_invoice_add",  # FIXED
                params={"shipment": obj.pk}
            )

    #from configuration.models import Choices  # adjust path if needed

    from configuration.models import ChoiceCategory

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        request.GET._mutable = True
        request.GET = request.GET.copy()

        # Correct dynamic status list
        from configuration.models import ChoiceCategory
        status_choices = Choice.objects.filter(
            category=ChoiceCategory.GENERAL_STATUS
        ).values_list("internal_value", "display_value")
        # print("FILTER USING ChoiceCategory.GENERAL_STATUS:", list(status_choices))
        #
        # print("===========================================================\n\n")
        extra_context["status_choices"] = list(status_choices)

        extra_context["vehicles"] = Vehicle.objects.only("id", "registration_number")
        extra_context["drivers"] = Driver.objects.only("id", "first_name", "last_name")

        return super().changelist_view(request, extra_context)

    def get_admin_url(self):
        from django.urls import reverse
        return reverse("admin:operations_shipment_change", args=[self.pk])

    def change_view(self, request, object_id, form_url='', extra_context=None):
        obj = self.get_object(request, object_id)

        # ---------------------------------------------------------
        # NEXT → Invoice
        # ---------------------------------------------------------
        invoice = obj.invoices.first()

        if invoice:
            next_step_url = reverse("admin:financial_invoice_change", args=[invoice.pk])
            next_step_label = "View Invoice"
        else:
            next_step_url = reverse("admin:financial_invoice_add") + f"?shipment={obj.pk}"
            next_step_label = "Create Invoice"

        # ---------------------------------------------------------
        # BACK → ConsignmentGroup
        # ---------------------------------------------------------
        if obj.consignment_group:
            back_step_url = reverse("admin:operations_consignmentgroup_change", args=[obj.consignment_group.pk])
            back_step_label = "Back to Consignment Group"
        else:
            back_step_url = reverse("admin:index")
            back_step_label = "Back to Dashboard"

        # ---------------------------------------------------------
        # STATUS BUTTON (View/Add)
        # ---------------------------------------------------------
        status = obj.status_logs.first()  # correct related_name

        if status:
            status_step_url = (
                    reverse("admin:operations_shipmentstatus_changelist")
                    + f"?shipment__id__exact={obj.pk}"
            )
            status_step_label = "View Status Updates"
        else:
            status_step_url = (
                    reverse("admin:operations_shipmentstatus_add")
                    + f"?shipment={obj.pk}"
            )
            status_step_label = "Add Status Update"

        # ---------------------------------------------------------
        # EXPENSE BUTTON (View/Add)
        # ---------------------------------------------------------
        expense = obj.expenses.first()  # correct related_name

        if expense:
            expense_step_url = (
                    reverse("admin:operations_shipmentexpense_changelist")
                    + f"?shipment__id__exact={obj.pk}"
            )
            expense_step_label = "View Expenses"
        else:
            expense_step_url = (
                    reverse("admin:operations_shipmentexpense_add")
                    + f"?shipment={obj.pk}"
            )
            expense_step_label = "Add Expense"

        # ---------------------------------------------------------
        # INJECT INTO TEMPLATE CONTEXT
        # ---------------------------------------------------------
        extra_context = extra_context or {}
        extra_context.update({
            "next_step_url": next_step_url,
            "next_step_label": next_step_label,
            "back_step_url": back_step_url,
            "back_step_label": back_step_label,

            "status_step_url": status_step_url,
            "status_step_label": status_step_label,

            "expense_step_url": expense_step_url,
            "expense_step_label": expense_step_label,
        })

        return super().change_view(request, object_id, form_url, extra_context)

    @admin.display(description="Status", ordering=None)
    def status_step_list_button(self, obj):
        count = obj.status_logs.count()

        if count == 0:
            return self.nav_button(
                "Add",
                "admin:operations_shipmentstatus_add",
                params={"shipment": obj.pk}
            )

        # Compact layout: Add More (top) + View All (bottom)
        return format_html(
            '<div style="display:flex; flex-direction:column; gap:2px;">'
            '   <div>{}</div>'
            '   <div>{}</div>'
            '</div>',
            self.nav_button(
                "Add More",
                "admin:operations_shipmentstatus_add",
                params={"shipment": obj.pk}
            ),
            self.nav_button(
                "View All",
                "admin:operations_shipmentstatus_changelist",
                params={"shipment__id__exact": obj.pk}
            )
        )

    @admin.display(description="Expenses", ordering=None)
    def expense_step_list_button(self, obj):
        count = obj.expenses.count()

        if count == 0:
            return self.nav_button(
                "Add",
                "admin:operations_shipmentexpense_add",
                params={"shipment": obj.pk}
            )

        return format_html(
            '<div style="display:flex; flex-direction:column; gap:2px;">'
            '   <div>{}</div>'
            '   <div>{}</div>'
            '</div>',
            self.nav_button(
                "Add More",
                "admin:operations_shipmentexpense_add",
                params={"shipment": obj.pk}
            ),
            self.nav_button(
                "View All",
                "admin:operations_shipmentexpense_changelist",
                params={"shipment__id__exact": obj.pk}
            )
        )

    @admin.display(description="Latest Status")
    def latest_status(self, obj):
        latest = obj.status_logs.order_by("-effective_date").first()
        if latest:
            return latest.status.display_value or latest.status.name
        return "-"

    @admin.display(description="Latest Status")
    def latest_status_badge(self, obj):
        latest = obj.status_logs.order_by("-effective_date").first()
        if not latest:
            return "-"

        # Correct values
        internal = latest.status.internal_value or ""
        display = latest.status.display_value or internal

        # Normalize key for color lookup
        key = internal.replace(" ", "").replace("_", "").upper()

        COLORS = {
            "00PLACED": "gray",
            "01LOADING": "yellow",
            "02WAITINGFORLORRYRECEIPT": "orange",
            "03DEPARTURE": "orange",
            "04INTRANSIT": "blue",
            "05REACHED": "cyan",
            "06UNLOADED": "purple",
            "07COMPLETED": "lime",
            "BREAKDOWN": "red",
            "CANCELLED": "red",
            "WORKSHOPVISIT": "white",
        }

        color = COLORS.get(key, "white")

        url = reverse("admin:operations_shipmentstatus_change", args=[latest.pk])

        return format_html(
            '<a href="{}" style="color:{}; font-weight:bold; text-decoration:none;">{}</a>',
            url,
            color,
            display
        )

    calculate_totals.short_description = "Recalculate totals for selected shipments"
    calculate_distances.short_description = "Calculate distances from odometer readings"

    def get_search_results(self, request, queryset, search_term):
        return queryset, False  # leave unchanged

    # def get_list_filter(self, request):
    #     return super().get_list_filter(request) + ('status',)


from django.urls import path
from django.http import JsonResponse
from django.db.models import Q


# Keep existing admin classes for supporting models
@admin.register(ShipmentExpense)
# class ShipmentExpenseAdmin(admin.ModelAdmin):
class ShipmentExpenseAdmin(NavigationButtonMixin, admin.ModelAdmin):
    """Admin for managing shipment expenses"""
    form = ShipmentExpenseForm
    autocomplete_fields = ("shipment", "expense_type",)
    fieldsets = (
        ('Expense Details', {
            'fields': ('shipment', 'expense_by_combined',
                       'expense_type', 'amount', 'expense_date')
        }),
        ('Description & Documentation', {
            'fields': ('description', 'shipment_expense_document'),
            'classes': ('collapse',),
        }),
    )

    # Exclude the GenericForeignKey fields from admin to prevent conflicts
    exclude = ['content_type', 'object_id', 'expense_by']

    list_display = ('shipment_display', 'expense_type', 'amount', 'expense_date', 'who_display')

    # list_filter = ('expense_type', 'expense_date', 'content_type')
    search_fields = ('shipment__vehicle__registration_number', 'description')
    date_hierarchy = 'expense_date'

    class Media:
        css = {
            "all": ("operations/css/admin_select2_fix.css",)
        }
        js = ("operations/js/expense_by_select2.js",)

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "expense-by-autocomplete/",
                self.admin_site.admin_view(self.expense_by_autocomplete),
                name="operations_expense_by_autocomplete",
            )
        ]
        return custom + urls

    def expense_by_autocomplete(self, request):
        term = (request.GET.get("term") or request.GET.get("q") or "").strip()

        results = []

        # Drivers
        drivers = Driver.objects.all()
        if term:
            drivers = drivers.filter(
                Q(first_name__icontains=term) |
                Q(last_name__icontains=term) |
                Q(phone_number__icontains=term)
            )
        drivers = drivers.order_by("first_name", "last_name")[:20]

        for d in drivers:
            label = f"Driver: {d.first_name} {d.last_name or ''}".strip()
            results.append({"id": f"driver_{d.id}", "text": label})

        # Owners
        owners = Organization.objects.filter(organization_type__internal_value="OWNER")
        if term:
            owners = owners.filter(organization_name__icontains=term)
        owners = owners.order_by("organization_name")[:20]

        for o in owners:
            results.append({"id": f"owner_{o.id}", "text": f"Owner: {o.organization_name}"})

        return JsonResponse({"results": results})

    def shipment_display(self, obj):
        """Display shipment with vehicle info"""
        if obj.shipment and hasattr(obj.shipment, 'vehicle'):
            return f"{obj.shipment} ({obj.shipment.vehicle.registration_number})"
        return str(obj.shipment) if obj.shipment else "No Shipment"

    def who_display(self, obj):
        """Display who made the expense"""
        content_type_obj_model = obj.content_type.model.capitalize()
        if not obj.content_type_id:
            return "-"
        try:
            target = obj.expense_by
            if hasattr(target, 'first_name'):  # Driver
                phone = getattr(target, 'phone_number', '')
                phone_display = f" ({phone})" if phone else ""
                last_name = target.last_name or ""
                full_name = f"{target.first_name} {last_name}".strip()
                return f"{content_type_obj_model}: {full_name}{phone_display}"
            else:  # Owner organization or other
                if obj.content_type.model == 'organization':
                    name = getattr(target, 'organization_name', str(target))
                    return f"Owner: {name}"
                name = getattr(target, "name", None) or str(target)
                return f"{content_type_obj_model}: {name}"
        except Exception:
            return f"{content_type_obj_model} (ID: {obj.object_id})"

    def related_advance_info(self, obj):
        """Show related driver advance information"""
        if not obj.shipment or not obj.expense_by:
            return "-"

        # Check if this is a driver expense
        if obj.content_type and obj.content_type.model == 'driver':
            try:
                from operations.models import DriverAdvance
                advance = DriverAdvance.objects.filter(
                    driver_id=obj.object_id,
                    shipment=obj.shipment
                ).first()

                if advance:
                    balance = advance.remaining_balance()
                    status = "✅ Settled" if advance.is_settled else "⏳ Pending"
                    return f"Advance: ₹{advance.amount} | Balance: ₹{balance} | {status}"
                else:
                    return "No advance found"
            except Exception:
                return "Error checking advance"
        return "-"

    change_form_template = "admin/consignment/change_form.html"

    class Media:
        css = {
            "all": ("operations/css/navigation_buttons.css",)
        }

    def change_view(self, request, object_id, form_url='', extra_context=None):
        obj = self.get_object(request, object_id)

        # BACK → Shipment (safe)
        if obj.shipment:
            back_step_url = reverse("admin:operations_shipment_change", args=[obj.shipment.pk])
            back_step_label = "Back to Shipment"
        else:
            back_step_url = reverse("admin:operations_shipment_changelist")
            back_step_label = "Back to Shipments"

        extra_context = extra_context or {}
        extra_context.update({
            "back_step_url": back_step_url,
            "back_step_label": back_step_label,
            "next_step_url": None,
            "next_step_label": None,
        })

        return super().change_view(request, object_id, form_url, extra_context)

    # Field descriptions
    shipment_display.short_description = "Shipment"
    who_display.short_description = "Expense By"
    who_display.admin_order_field = "object_id"
    # related_advance_info.short_description = "Related Advance"


@admin.register(ShipmentStatus)
# class ShipmentStatusAdmin(admin.ModelAdmin):
class ShipmentStatusAdmin(NavigationButtonMixin, admin.ModelAdmin):
    """Admin for shipment status tracking"""
    fieldsets = (
        ('Status Information', {
            'fields': ('shipment', 'status', 'shipment_doc_type', 'shipment_document', 'effective_date', 'notes')
        }),
        ('Update Details', {
            'fields': ('updated_by',),
            'classes': ('collapse',),
            'description': 'Who updated the status and additional notes.'
        })
    )

    list_display = ('shipment', 'status', 'effective_date', 'updated_by')
    # list_filter = ('status', 'effective_date')
    search_fields = ('shipment__shipment_id', 'shipment__vehicle__registration_number', 'updated_by')
    date_hierarchy = 'effective_date'
    # readonly_fields = ('effective_date',)
    autocomplete_fields = ('shipment', 'status',)

    change_form_template = "admin/consignment/change_form.html"

    class Media:
        css = {
            "all": ("operations/css/navigation_buttons.css",)
        }

    def save_model(self, request, obj, form, change):
        """
        Auto-update Shipment fields based on status.
        """
        super().save_model(request, obj, form, change)

        shipment = obj.shipment
        if not shipment:
            return

        # Normalize status
        raw_status = ""
        if obj.status:
            raw_status = obj.status.internal_value or obj.status.name or ""

        status_name = raw_status.replace(" ", "").replace("-", "").lower()

        # -----------------------------
        # 1. Departure → actual_departure + odometer_start (optional)
        # -----------------------------
        if status_name in ["03_departure", "departure", "departed"]:
            shipment.actual_departure = obj.effective_date

            # If notes contains odometer reading
            if obj.notes:
                try:
                    shipment.odometer_start = float(obj.notes.strip())
                except:
                    pass

        # -----------------------------
        # 2. Reached → actual_arrival + odometer_start
        # -----------------------------
        elif status_name in ["04_reached", "reached", "arrived"]:
            shipment.actual_arrival = obj.effective_date

            if obj.notes:
                try:
                    shipment.odometer_start = float(obj.notes.strip())
                except:
                    pass

        # -----------------------------
        # 3. Completed → odometer_end
        # -----------------------------
        elif status_name in ["07_completed", "completed", "finished"]:
            if obj.notes:
                try:
                    shipment.odometer_end = float(obj.notes.strip())
                except:
                    pass

        # -----------------------------
        # 4. Auto-calc total distance
        # -----------------------------
        if shipment.odometer_start and shipment.odometer_end:
            shipment.total_distance = shipment.odometer_end - shipment.odometer_start

        shipment.save()

    def change_view(self, request, object_id, form_url='', extra_context=None):
        obj = self.get_object(request, object_id)

        # BACK → Shipment (safe)
        if obj.shipment:
            back_step_url = reverse("admin:operations_shipment_change", args=[obj.shipment.pk])
            back_step_label = "Back to Shipment"
        else:
            back_step_url = reverse("admin:operations_shipment_changelist")
            back_step_label = "Back to Shipments"

        extra_context = extra_context or {}
        extra_context.update({
            "back_step_url": back_step_url,
            "back_step_label": back_step_label,
            "next_step_url": None,
            "next_step_label": None,
        })

        return super().change_view(request, object_id, form_url, extra_context)


def _admin_url_for_obj(obj):
    try:
        return reverse(f'admin:{obj._meta.app_label}_{obj._meta.model_name}_change', args=[obj.pk])
    except Exception:
        return None


@admin.register(DriverAdvance)
class DriverAdvanceAdmin(admin.ModelAdmin):
    form = DriverAdvanceAdminForm
    exclude = ('content_type', 'object_id')  # keep GFK hidden; the form sets it
    autocomplete_fields = ('driver', 'shipment')

    # ---------- UI layout ----------
    fieldsets = (
        ('Advance Details', {
            'fields': (
                'driver',
                ('related_type', 'owner_ref'),  # ONE owner control
                'shipment',  # (optional) still available but not required
                'amount', 'description',
            )
        }),
        ('Computed (Read-only)', {
            'fields': (
                'date',
                'total_expenses',
                'carried_forward',
                'is_settled',
                'settlement_status_display',
                'advance_by_display',
            )
        }),
        ('Driver Ledger', {
            'fields': ('driver_full_ledger',),
            'classes': ('collapse',),
            'description': 'Unified ledger showing all advances (credit) and expenses (debit).'
        }),

    )

    # ---------- list view ----------
    list_display = (
        'driver_name',
        'amount',  # positive = advance, negative = maintenance deduction (or other deduction)
        'total_expenses',  # cumulative expenses for this shipment+driver (unchanged)
        'carried_forward',
        'settlement_status_display',
        'date',
    )
    # list_filter = ('is_settled', 'date', 'driver')
    search_fields = ('driver__first_name', 'driver__last_name', 'description')
    readonly_fields = (
        'date', 'total_expenses', 'carried_forward', 'is_settled',
        'settlement_status_display', 'advance_by_display'
        , 'driver_full_ledger',

    )

    # 'driver_financial_summary', 'driver_ledger_preview', 'driver_expense_summary'
    date_hierarchy = 'date'
    actions = ['recompute_carry_forward_for_selected_drivers']

    # ---------- presenters ----------
    def driver_name(self, obj):
        phone = getattr(obj.driver, 'phone_number', '')
        phone_display = f" ({phone})" if phone else ""
        last = getattr(obj.driver, 'last_name', '') or ''
        return f"{obj.driver.first_name} {last}{phone_display}".strip()

    driver_name.short_description = "Driver"

    def shipment_display(self, obj):
        return str(obj.shipment) if obj.shipment else "—"

    shipment_display.short_description = "Shipment"

    def settlement_status_display(self, obj):
        return "✅ Settled" if obj.is_settled else "⏳ Pending"

    settlement_status_display.short_description = "Status"

    def advance_by_display(self, obj):
        if not obj.content_type_id or not obj.object_id:
            return "—"
        try:
            model = obj.content_type.model_class()
            target = model.objects.get(pk=obj.object_id)
        except Exception:
            return f"{obj.content_type.app_label}.{obj.content_type.model}#{obj.object_id}"
        url = _admin_url_for_obj(target)
        return format_html('<a href="{}">{}</a>', url, str(target)) if url else str(target)

    advance_by_display.short_description = "Advance By"

    # ---------- summary + ledger ----------
    def _driver_totals(self, driver_id):
        """
        Shipment-independent financial totals for a driver across ALL DriverAdvance rows:
          - total_advances: sum of all positive amounts
          - total_deductions: sum(abs(negative amounts)) – e.g. maintenance deductions
          - net_advance: total_advances - total_deductions
          - current_carried_forward: carried_forward from the last row (after recompute)
        """
        qs = DriverAdvance.objects.filter(driver_id=driver_id).order_by('date', 'id')

        # sums
        sums = qs.aggregate(all_sum=Sum('amount'))
        all_sum = sums['all_sum'] or Decimal('0')

        positives = qs.filter(amount__gt=0).aggregate(s=Sum('amount'))['s'] or Decimal('0')
        negatives = qs.filter(amount__lt=0).aggregate(s=Sum('amount'))['s'] or Decimal('0')  # negative value
        total_advances = positives
        total_deductions = -negatives  # make positive number
        net_advance = total_advances - total_deductions

        # last carried_forward after recompute chain (if nothing exists -> 0)
        last_cf = qs.last().carried_forward if qs.exists() else Decimal('0')

        return {
            'total_advances': total_advances,
            'total_deductions': total_deductions,
            'net_advance': net_advance,
            'sum_all_rows': all_sum,
            'current_carried_forward': last_cf,
            'count': qs.count(),
        }

    def driver_financial_summary(self, obj):
        """
        Compact HTML box summarizing totals for this driver across all rows (shipment-agnostic).
        """
        t = self._driver_totals(obj.driver_id)
        html = f"""
        <div style="font-family:Inter,system-ui,-apple-system,Segoe UI,Roboto,Arial;line-height:1.5">
          <p><strong>Total Advances (₹):</strong> {t['total_advances']}</p>
          <p><strong>Total Deductions (₹):</strong> {t['total_deductions']}</p>
          <p><strong>Net Advance (₹):</strong> {t['net_advance']}</p>
          <p><strong>Current Carried Forward (₹):</strong> {t['current_carried_forward']}</p>
          <p style="color:#666">({t['count']} rows total; sum of all rows = {t['sum_all_rows']})</p>
        </div>
        """
        return format_html(html)

    driver_financial_summary.short_description = "Driver Financial Summary"

    def driver_ledger_preview(self, obj):
        """
        Show last 10 DriverAdvance rows for this driver with running CF, newest first.
        Negative amounts show as “maintenance deduction” style.
        """
        rows = (
            DriverAdvance.objects
            .filter(driver_id=obj.driver_id)
            .order_by('-date', '-id')[:10]
            .values('id', 'date', 'amount', 'carried_forward', 'description')
        )
        if not rows:
            return format_html("<em>No ledger entries</em>")

        parts = [
            '<table style="border-collapse:collapse;width:100%">',
            '<thead><tr>'
            '<th style="text-align:left;border-bottom:1px solid #ddd;padding:4px">ID</th>'
            '<th style="text-align:left;border-bottom:1px solid #ddd;padding:4px">Date</th>'
            '<th style="text-align:right;border-bottom:1px solid #ddd;padding:4px">Amount (₹)</th>'
            '<th style="text-align:right;border-bottom:1px solid #ddd;padding:4px">Carried Fwd (₹)</th>'
            '<th style="text-align:left;border-bottom:1px solid #ddd;padding:4px">Description</th>'
            '</tr></thead><tbody>'
        ]
        for r in rows:
            amt = Decimal(r['amount'] or 0)
            amt_style = 'color:#c62828' if amt < 0 else 'color:#1b5e20'
            parts.append(
                f"<tr>"
                f"<td style='padding:4px'>{r['id']}</td>"
                f"<td style='padding:4px'>{r['date']}</td>"
                f"<td style='padding:4px;text-align:right;{amt_style}'>{amt}</td>"
                f"<td style='padding:4px;text-align:right'>{r['carried_forward']}</td>"
                f"<td style='padding:4px'>{(r['description'] or '').replace('<', '&lt;').replace('>', '&gt;')}</td>"
                f"</tr>"
            )
        parts.append("</tbody></table>")
        return format_html("".join(parts))

    driver_ledger_preview.short_description = "Recent Ledger (10)"

    # ---------- data/queries ----------
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('driver', 'content_type', 'shipment')

    # ---------- actions ----------
    @admin.action(description="Recompute carry-forward for selected drivers")
    def recompute_carry_forward_for_selected_drivers(self, request, queryset):
        """
        For every selected row, recompute the *entire* chain for that driver once.
        This catches maintenance deductions (negative rows) and normal advances.
        """
        from .models import DriverAdvance as DA
        driver_ids = set(queryset.values_list('driver_id', flat=True))
        total_updated = 0
        for did in driver_ids:
            chain = DA.objects.filter(driver_id=did).order_by('date', 'id')
            for adv in chain:
                adv.settle_and_carry_forward()
                total_updated += 1
        messages.success(request,
                         f"Recomputed carry-forward for {len(driver_ids)} driver(s), updated {total_updated} rows.")

    def driver_expense_summary(self, obj):
        """
        Show all ShipmentExpense rows for this driver across all shipments.
        """
        from operations.models import ShipmentExpense
        from django.contrib.contenttypes.models import ContentType

        driver_ct = ContentType.objects.get_for_model(obj.driver)

        expenses = (
            ShipmentExpense.objects
            .filter(content_type=driver_ct, object_id=obj.driver_id)
            .select_related("shipment")
            .order_by("-expense_date", "-id")
            .values("id", "expense_type__display_value", "amount", "expense_date", "description",
                    "shipment__shipment_id")
        )

        if not expenses:
            return format_html("<em>No expenses recorded for this driver.</em>")

        parts = [
            '<table style="border-collapse:collapse;width:100%">',
            '<thead><tr>'
            '<th style="text-align:left;border-bottom:1px solid #ddd;padding:4px">ID</th>'
            '<th style="text-align:left;border-bottom:1px solid #ddd;padding:4px">Shipment</th>'
            '<th style="text-align:left;border-bottom:1px solid #ddd;padding:4px">Type</th>'
            '<th style="text-align:right;border-bottom:1px solid #ddd;padding:4px">Amount (₹)</th>'
            '<th style="text-align:left;border-bottom:1px solid #ddd;padding:4px">Date</th>'
            '<th style="text-align:left;border-bottom:1px solid #ddd;padding:4px">Description</th>'
            '</tr></thead><tbody>'
        ]

        for e in expenses:
            parts.append(
                f"<tr>"
                f"<td style='padding:4px'>{e['id']}</td>"
                f"<td style='padding:4px'>{e['shipment__shipment_id'] or '—'}</td>"
                f"<td style='padding:4px'>{e['expense_type__display_value'] or '—'}</td>"
                f"<td style='padding:4px;text-align:right'>{e['amount']}</td>"
                f"<td style='padding:4px'>{e['expense_date']}</td>"
                f"<td style='padding:4px'>{(e['description'] or '').replace('<', '&lt;').replace('>', '&gt;')}</td>"
                f"</tr>"
            )

        parts.append("</tbody></table>")
        return format_html("".join(parts))

    driver_expense_summary.short_description = "Driver Expense Summary"

    def driver_full_ledger(self, obj):
        request = getattr(self, "_request", None)

        ledger = self._build_driver_ledger(obj.driver)

        opening_balance = ledger["opening_balance"]
        closing_balance = ledger["closing_balance"]
        rows = ledger["rows"]

        # -----------------------------
        # Build filter + export UI
        # -----------------------------
        export_url = reverse("admin:driver_ledger_excel", args=[obj.driver_id])

        filter_html = ""
        if request:
            from_val = request.GET.get("from", "")
            to_val = request.GET.get("to", "")

            request = getattr(self, "_request", None)
            # current_url = request.path if request else ""
            current_url = self._request.path

            filter_html = f"""
            <form method="get" action="{current_url}" style="display:inline-block">



            <div style='margin-bottom:10px'>
                <a href="{export_url}" class="button" style="margin-right:10px">Download Ledger (Excel)</a>

                <form method="get" action="{current_url}" style="display:inline-block">
                    <label>From: <input type="date" name="from" value="{request.GET.get('from', '')}"></label>
                    <label>To: <input type="date" name="to" value="{request.GET.get('to', '')}"></label>
                    <button type="submit" class="button">Filter</button>
                </form>
            </div>
            """

        # -----------------------------
        # Build table
        # -----------------------------
        parts = [
            "<div style='margin-bottom:10px'>"
            f"<strong>Opening Balance:</strong> ₹ {opening_balance}<br>"
            f"<strong>Closing Balance:</strong> ₹ {closing_balance}"
            "</div>",

            '<table style="border-collapse:collapse;width:100%">',
            '<thead><tr>'
            '<th style="padding:4px">Date</th>'
            '<th style="padding:4px">Type</th>'
            '<th style="padding:4px">Shipment</th>'
            '<th style="padding:4px;text-align:right">Debit (₹)</th>'
            '<th style="padding:4px;text-align:right">Credit (₹)</th>'
            '<th style="padding:4px;text-align:right">Running Balance (₹)</th>'
            '<th style="padding:4px">Description</th>'
            '</tr></thead><tbody>'
        ]

        for row in rows:
            parts.append(
                f"<tr>"
                f"<td style='padding:4px'>{row['date']}</td>"
                f"<td style='padding:4px'>{row['type']}</td>"
                f"<td style='padding:4px'>{row['shipment'] or '—'}</td>"
                f"<td style='padding:4px;text-align:right;color:#c62828'>{row['debit'] or ''}</td>"
                f"<td style='padding:4px;text-align:right;color:#1b5e20'>{row['credit'] or ''}</td>"
                f"<td style='padding:4px;text-align:right'>{row['running_balance']}</td>"
                f"<td style='padding:4px'>{row['description']}</td>"
                f"</tr>"
            )

        parts.append("</tbody></table>")

        return format_html(filter_html + "".join(parts))

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "export-ledger/<int:driver_id>/",
                self.admin_site.admin_view(self.export_ledger_excel),
                name="driver_ledger_excel",
            )
        ]
        return custom + urls

    def export_ledger_excel(self, request, driver_id):
        """
        Export full driver ledger (advances + expenses) to Excel.
        """
        driver = Driver.objects.get(pk=driver_id)

        # Build ledger using the same logic as driver_full_ledger()
        ledger = self._build_driver_ledger(driver)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Driver Ledger"

        headers = [
            "Date", "Type", "Shipment", "Debit (₹)", "Credit (₹)",
            "Running Balance (₹)", "Description"
        ]
        ws.append(headers)

        for row in ledger["rows"]:
            ws.append([
                row["date"],
                row["type"],
                row["shipment"],
                row["debit"],
                row["credit"],
                row["running_balance"],
                row["description"],
            ])

        # Auto-fit columns
        for col in ws.columns:
            length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = length + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="driver_ledger_{driver_id}.xlsx"'
        wb.save(response)
        return response

    def _build_driver_ledger(self, driver):
        from operations.models import ShipmentExpense
        from django.contrib.contenttypes.models import ContentType
        from decimal import Decimal

        driver_ct = ContentType.objects.get_for_model(driver)

        ledger = []

        # ---------------------------------------------------------
        # 1) DRIVER ADVANCES (CREDIT)
        # ---------------------------------------------------------
        advances = DriverAdvance.objects.filter(driver=driver).values(
            "date", "amount", "description", "shipment__shipment_id"
        )

        for a in advances:
            ledger.append({
                "date": a["date"],
                "type": "Advance",
                "shipment": a["shipment__shipment_id"],
                "debit": Decimal("0"),
                "credit": Decimal(a["amount"] or 0),
                "description": a["description"] or "",
            })

        # ---------------------------------------------------------
        # 2) SHIPMENT EXPENSES (DEBIT)
        # ---------------------------------------------------------
        expenses = ShipmentExpense.objects.filter(
            content_type=driver_ct, object_id=driver.id
        ).values(
            "expense_date", "amount", "description", "shipment__shipment_id",
            "expense_type__display_value"
        )

        for e in expenses:
            ledger.append({
                "date": e["expense_date"],
                "type": e["expense_type__display_value"] or "Shipment Expense",
                "shipment": e["shipment__shipment_id"],
                "debit": Decimal(e["amount"] or 0),
                "credit": Decimal("0"),
                "description": e["description"] or "",
            })

        # ---------------------------------------------------------
        # SORT + RUNNING BALANCE
        # ---------------------------------------------------------
        ledger.sort(key=lambda x: (x["date"] or "", x["type"]))

        running = Decimal("0")
        for row in ledger:
            running += row["credit"]
            running -= row["debit"]
            row["running_balance"] = running

        return {
            "rows": ledger,
            "opening_balance": Decimal("0"),
            "closing_balance": running,
        }

    def change_view(self, request, object_id, form_url='', extra_context=None):
        # Preserve GET parameters so Django Admin does NOT redirect
        request.GET._mutable = True

        # Pass request to ledger
        extra_context = extra_context or {}
        extra_context["request"] = request

        return super().change_view(request, object_id, form_url, extra_context)

    def get_form(self, request, obj=None, **kwargs):
        self._request = request
        return super().get_form(request, obj, **kwargs)

    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        # Preserve GET parameters so Django Admin does NOT redirect
        request.GET._mutable = True

        # Pass request to ledger
        context["request"] = request

        return super().render_change_form(request, context, add, change, form_url, obj)


@admin.register(Diesel)
class DieselAdmin(admin.ModelAdmin):
    # ✅ Table columns in admin list page
    list_display = (
        "date",
        "vehicle",
        "driver",
        "quantity",
        "price_per_ltr",
        "total_price",
        "payment",
        "payment_mode",
        "driver_taken_cash",
        "full_km",
        "mileage",
        "rs_per_km",
        "created_at",
    )

    # ✅ Sidebar filters
    # list_filter = (
    #     "payment_mode",
    #     "date",
    #     "vehicle",
    #     "driver",
    # )

    # ✅ Search box (adjust field names to your Vehicle/Driver models)
    search_fields = (
        "vehicle__registration_number",
        "driver__first_name",
        "driver__last_name",
    )

    # ✅ Default ordering
    ordering = ("-date", "-id")

    # ✅ Better UX for FK dropdowns (faster than huge dropdown)
    autocomplete_fields = ("vehicle", "driver", "location",)

    # ✅ Use raw_id_fields if you prefer (comment above and uncomment below)
    # raw_id_fields = ("vehicle", "driver")

    # ✅ Readonly calculated fields (so user can’t edit manually)
    readonly_fields = (
        "total_price",
        "mileage",
        "rs_per_km",
        "created_at",
        "payment",
    )

    # ✅ Form layout in admin detail page
    fieldsets = (
        ("Basic Info", {
            "fields": ("vehicle", "driver", "date", "location")
        }),
        ("Diesel Details", {
            "fields": ("price_per_ltr", "quantity", "total_price","driver_taken_cash")
        }),
        ("Payment", {
            "fields": ("payment", "payment_mode", )
        }),
        ("Odometer & Calculations", {
            "fields": ("full_km", "mileage", "rs_per_km")
        }),

        ("Documents", {
            "fields": ("upload_doc","description")
        }),
        ("System", {
            "fields": ("created_at",)
        }),
    )

    # ✅ Nice admin list page improvements
    list_per_page = 50
    save_on_top = True

    # Optional: show totals at top of changelist page
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        qs = self.get_queryset(request)

        totals = qs.aggregate(
            total_qty=Sum("quantity"),
            total_amount=Sum("total_price"),
            total_payment=Sum("payment"),
        )

        extra_context["diesel_totals"] = totals
        return super().changelist_view(request, extra_context=extra_context)

    from django.db.models import Q

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)

        if search_term:
            queryset = queryset.filter(
                Q(vehicle__registration_number__icontains=search_term)
                | Q(driver__first_name__icontains=search_term)
                | Q(driver__last_name__icontains=search_term)
                | Q(driver__middle_name__icontains=search_term)
            )

        return queryset, True

